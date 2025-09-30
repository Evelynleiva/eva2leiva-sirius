from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.core.exceptions import PermissionDenied
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from .models import Cliente, Servicio, Proyecto, Presupuesto, Incidencia, PerfilUsuario
from .forms import (
    ClienteForm, ServicioForm, ProyectoForm, PresupuestoForm, 
    IncidenciaForm, IncidenciaResolucionForm, CustomUserCreationForm, 
    PerfilUsuarioForm, ProyectoFiltroForm, IncidenciaFiltroForm
)

# Vista principal/home
def home(request):
    """Página principal con estadísticas generales"""
    context = {
        'total_proyectos': 0,
        'proyectos_activos': 0,
        'total_clientes': 0,
        'incidencias_abiertas': 0,
        'proyectos_recientes': [],
        'servicios': [],
    }
    
    if request.user.is_authenticated:
        context['total_proyectos'] = Proyecto.objects.count()
        context['proyectos_activos'] = Proyecto.objects.filter(estado='en_proceso').count()
        context['total_clientes'] = Cliente.objects.filter(activo=True).count()
        context['incidencias_abiertas'] = Incidencia.objects.filter(estado__in=['abierta', 'en_proceso']).count()
        context['proyectos_recientes'] = Proyecto.objects.all()[:5]
        context['servicios'] = Servicio.objects.filter(activo=True)[:4]
    
    return render(request, 'home.html', context)

# ============ VISTAS DE CLIENTES ============

@login_required
def cliente_lista(request):
    """Lista de clientes con paginación"""
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    paginator = Paginator(clientes, 10)
    page = request.GET.get('page')
    clientes = paginator.get_page(page)
    
    context = {'clientes': clientes}
    return render(request, 'clientes/lista.html', context)

@login_required
def cliente_crear(request):
    """Crear nuevo cliente"""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            messages.success(request, f'Cliente {cliente.nombre} creado exitosamente.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm()
    
    context = {'form': form, 'titulo': 'Nuevo Cliente'}
    return render(request, 'clientes/form.html', context)

@login_required
def cliente_editar(request, pk):
    """Editar cliente existente"""
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente {cliente.nombre} actualizado exitosamente.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm(instance=cliente)
    
    context = {'form': form, 'titulo': f'Editar Cliente: {cliente.nombre}', 'cliente': cliente}
    return render(request, 'clientes/form.html', context)

@login_required
@staff_member_required
def cliente_eliminar(request, pk):
    """Eliminar cliente (desactivar) - Solo admin"""
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = False
    cliente.save()
    messages.success(request, f'Cliente {cliente.nombre} eliminado exitosamente.')
    return redirect('cliente_lista')

# ============ VISTAS DE SERVICIOS ============

@login_required
def servicio_lista(request):
    """Lista de servicios"""
    servicios = Servicio.objects.filter(activo=True)
    context = {'servicios': servicios}
    return render(request, 'servicios/lista.html', context)

@login_required
def servicio_crear(request):
    """Crear nuevo servicio - Solo admin"""
    # Verificar si es admin
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para crear servicios.')
        return redirect('servicio_lista')
    
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            servicio = form.save()
            messages.success(request, f'Servicio {servicio.nombre} creado exitosamente.')
            return redirect('servicio_lista')
    else:
        form = ServicioForm()
    
    context = {'form': form, 'titulo': 'Nuevo Servicio'}
    return render(request, 'servicios/form.html', context)

@login_required
def servicio_editar(request, pk):
    """Editar servicio existente"""
    servicio = get_object_or_404(Servicio, pk=pk)
    
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, f'Servicio {servicio.nombre} actualizado exitosamente.')
            return redirect('servicio_lista')
    else:
        form = ServicioForm(instance=servicio)
    
    context = {'form': form, 'titulo': f'Editar Servicio: {servicio.nombre}', 'servicio': servicio}
    return render(request, 'servicios/form.html', context)

@login_required
@staff_member_required
def servicio_eliminar(request, pk):
    """Eliminar servicio (desactivar) - Solo admin"""
    servicio = get_object_or_404(Servicio, pk=pk)
    servicio.activo = False
    servicio.save()
    messages.success(request, f'Servicio {servicio.nombre} eliminado exitosamente.')
    return redirect('servicio_lista')

# ============ VISTAS DE PROYECTOS ============

@login_required
def proyecto_lista(request):
    """Lista de proyectos con filtros y paginación"""
    proyectos = Proyecto.objects.all()
    filtro_form = ProyectoFiltroForm(request.GET)
    
    # Aplicar filtros
    if filtro_form.is_valid():
        if filtro_form.cleaned_data['cliente']:
            proyectos = proyectos.filter(cliente=filtro_form.cleaned_data['cliente'])
        if filtro_form.cleaned_data['estado']:
            proyectos = proyectos.filter(estado=filtro_form.cleaned_data['estado'])
        if filtro_form.cleaned_data['prioridad']:
            proyectos = proyectos.filter(prioridad=filtro_form.cleaned_data['prioridad'])
        if filtro_form.cleaned_data['fecha_inicio']:
            proyectos = proyectos.filter(fecha_inicio__gte=filtro_form.cleaned_data['fecha_inicio'])
        if filtro_form.cleaned_data['fecha_fin']:
            proyectos = proyectos.filter(fecha_fin_estimada__lte=filtro_form.cleaned_data['fecha_fin'])
        if filtro_form.cleaned_data['responsable']:
            proyectos = proyectos.filter(responsable=filtro_form.cleaned_data['responsable'])
    
    # RESTRICCIÓN: Verificar permisos de usuario
    perfil_usuario = getattr(request.user, 'perfilusuario', None)
    if perfil_usuario and perfil_usuario.tipo_usuario == 'cliente':
        # Los clientes solo ven sus proyectos
        cliente = Cliente.objects.filter(email=request.user.email).first()
        if cliente:
            proyectos = proyectos.filter(cliente=cliente)
    elif not request.user.is_staff:
        # Usuarios normales (empleados, contratistas) solo ven proyectos donde son responsables
        proyectos = proyectos.filter(responsable=request.user)
    
    # Paginación
    paginator = Paginator(proyectos, 10)
    page = request.GET.get('page')
    proyectos = paginator.get_page(page)
    
    context = {
        'proyectos': proyectos,
        'filtro_form': filtro_form
    }
    return render(request, 'proyectos/lista.html', context)

@login_required
def proyecto_crear(request):
    """Crear nuevo proyecto"""
    if request.method == 'POST':
        form = ProyectoForm(request.POST, user=request.user)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.creado_por = request.user
            proyecto.save()
            form.save_m2m()  # Guardar relaciones ManyToMany
            messages.success(request, f'Proyecto {proyecto.nombre} creado exitosamente.')
            return redirect('proyecto_lista')
    else:
        form = ProyectoForm(user=request.user)
    
    context = {'form': form, 'titulo': 'Nuevo Proyecto'}
    return render(request, 'proyectos/form.html', context)

@login_required
def proyecto_detalle(request, pk):
    """Ver detalle del proyecto"""
    proyecto = get_object_or_404(Proyecto, pk=pk)
    presupuestos = proyecto.presupuestos.all()
    incidencias = proyecto.incidencias.all()[:5]  # Últimas 5 incidencias
    
    context = {
        'proyecto': proyecto,
        'presupuestos': presupuestos,
        'incidencias': incidencias
    }
    return render(request, 'proyectos/detalle.html', context)

@login_required
def proyecto_editar(request, pk):
    """Editar proyecto existente"""
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f'Proyecto {proyecto.nombre} actualizado exitosamente.')
            return redirect('proyecto_detalle', pk=proyecto.pk)
    else:
        form = ProyectoForm(instance=proyecto, user=request.user)
    
    context = {'form': form, 'titulo': f'Editar Proyecto: {proyecto.nombre}', 'proyecto': proyecto}
    return render(request, 'proyectos/form.html', context)

# ============ VISTAS DE PRESUPUESTOS ============

@login_required
def presupuesto_lista(request):
    """Lista de presupuestos"""
    presupuestos = Presupuesto.objects.all()
    
    # Filtrar por cliente si es necesario
    perfil_usuario = getattr(request.user, 'perfilusuario', None)
    if perfil_usuario and perfil_usuario.tipo_usuario == 'cliente':
        cliente = Cliente.objects.filter(email=request.user.email).first()
        if cliente:
            presupuestos = presupuestos.filter(cliente=cliente)
    
    paginator = Paginator(presupuestos, 10)
    page = request.GET.get('page')
    presupuestos = paginator.get_page(page)
    
    context = {'presupuestos': presupuestos}
    return render(request, 'presupuestos/lista.html', context)

@login_required
def presupuesto_crear(request):
    """Crear nuevo presupuesto"""
    if request.method == 'POST':
        form = PresupuestoForm(request.POST)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.creado_por = request.user
            presupuesto.save()
            messages.success(request, f'Presupuesto {presupuesto.numero_presupuesto} creado exitosamente.')
            return redirect('presupuesto_lista')
    else:
        form = PresupuestoForm()
    
    context = {'form': form, 'titulo': 'Nuevo Presupuesto'}
    return render(request, 'presupuestos/form.html', context)

@login_required
def presupuesto_detalle(request, pk):
    """Ver detalle del presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    context = {'presupuesto': presupuesto}
    return render(request, 'presupuestos/detalle.html', context)

# ============ VISTAS DE INCIDENCIAS ============

@login_required
def incidencia_lista(request):
    """Lista de incidencias con filtros"""
    incidencias = Incidencia.objects.all()
    filtro_form = IncidenciaFiltroForm(request.GET)
    
    # Aplicar filtros
    if filtro_form.is_valid():
        if filtro_form.cleaned_data['proyecto']:
            incidencias = incidencias.filter(proyecto=filtro_form.cleaned_data['proyecto'])
        if filtro_form.cleaned_data['tipo_incidencia']:
            incidencias = incidencias.filter(tipo_incidencia=filtro_form.cleaned_data['tipo_incidencia'])
        if filtro_form.cleaned_data['estado']:
            incidencias = incidencias.filter(estado=filtro_form.cleaned_data['estado'])
        if filtro_form.cleaned_data['prioridad']:
            incidencias = incidencias.filter(prioridad=filtro_form.cleaned_data['prioridad'])
    
    paginator = Paginator(incidencias, 10)
    page = request.GET.get('page')
    incidencias = paginator.get_page(page)
    
    context = {
        'incidencias': incidencias,
        'filtro_form': filtro_form
    }
    return render(request, 'incidencias/lista.html', context)

@login_required
def incidencia_crear(request):
    """Crear nueva incidencia"""
    if request.method == 'POST':
        form = IncidenciaForm(request.POST, request.FILES)
        if form.is_valid():
            incidencia = form.save(commit=False)
            incidencia.reportado_por = request.user
            incidencia.save()
            messages.success(request, f'Incidencia {incidencia.titulo} creada exitosamente.')
            return redirect('incidencia_lista')
    else:
        form = IncidenciaForm()
    
    context = {'form': form, 'titulo': 'Nueva Incidencia'}
    return render(request, 'incidencias/form.html', context)

@login_required
def incidencia_resolver(request, pk):
    """Resolver incidencia"""
    incidencia = get_object_or_404(Incidencia, pk=pk)
    
    if request.method == 'POST':
        form = IncidenciaResolucionForm(request.POST, instance=incidencia)
        if form.is_valid():
            incidencia = form.save(commit=False)
            if incidencia.estado in ['resuelta', 'cerrada']:
                incidencia.fecha_resolucion = timezone.now()
            incidencia.save()
            messages.success(request, f'Incidencia {incidencia.titulo} actualizada.')
            return redirect('incidencia_lista')
    else:
        form = IncidenciaResolucionForm(instance=incidencia)
    
    context = {'form': form, 'incidencia': incidencia, 'titulo': 'Resolver Incidencia'}
    return render(request, 'incidencias/resolver.html', context)

# ============ VISTAS DE AUTENTICACIÓN ============

def registro(request):
    """Registro de nuevos usuarios"""
    if request.method == 'POST':
        user_form = CustomUserCreationForm(request.POST)
        perfil_form = PerfilUsuarioForm(request.POST, request.FILES)
        
        if user_form.is_valid() and perfil_form.is_valid():
            user = user_form.save()
            perfil = perfil_form.save(commit=False)
            perfil.user = user
            perfil.save()
            
            username = user_form.cleaned_data.get('username')
            messages.success(request, f'Cuenta creada para {username}!')
            return redirect('login')
    else:
        user_form = CustomUserCreationForm()
        perfil_form = PerfilUsuarioForm()
    
    context = {
        'user_form': user_form,
        'perfil_form': perfil_form
    }
    return render(request, 'registration/registro.html', context)

@login_required
def perfil(request):
    """Ver y editar perfil de usuario"""
    perfil_usuario, created = PerfilUsuario.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = PerfilUsuarioForm(request.POST, request.FILES, instance=perfil_usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('perfil')
    else:
        form = PerfilUsuarioForm(instance=perfil_usuario)
    
    context = {'form': form, 'perfil': perfil_usuario}
    return render(request, 'registration/perfil.html', context)

# ============ VISTA DE EXPORTACIÓN A EXCEL ============

@login_required
def exportar_proyectos_excel(request):
    """Exportar proyectos filtrados a Excel"""
    proyectos = Proyecto.objects.all()
    filtro_form = ProyectoFiltroForm(request.GET)
    
    # Aplicar los mismos filtros que en la lista
    if filtro_form.is_valid():
        if filtro_form.cleaned_data['cliente']:
            proyectos = proyectos.filter(cliente=filtro_form.cleaned_data['cliente'])
        if filtro_form.cleaned_data['estado']:
            proyectos = proyectos.filter(estado=filtro_form.cleaned_data['estado'])
        if filtro_form.cleaned_data['prioridad']:
            proyectos = proyectos.filter(prioridad=filtro_form.cleaned_data['prioridad'])
        if filtro_form.cleaned_data['fecha_inicio']:
            proyectos = proyectos.filter(fecha_inicio__gte=filtro_form.cleaned_data['fecha_inicio'])
        if filtro_form.cleaned_data['fecha_fin']:
            proyectos = proyectos.filter(fecha_fin_estimada__lte=filtro_form.cleaned_data['fecha_fin'])
        if filtro_form.cleaned_data['responsable']:
            proyectos = proyectos.filter(responsable=filtro_form.cleaned_data['responsable'])
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos Sirius"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Cliente', 'Estado', 'Prioridad', 
        'Fecha Inicio', 'Fecha Fin Est.', 'Presupuesto', 'Responsable'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, proyecto in enumerate(proyectos, 2):
        ws.cell(row=row, column=1, value=proyecto.id)
        ws.cell(row=row, column=2, value=proyecto.nombre)
        ws.cell(row=row, column=3, value=str(proyecto.cliente))
        ws.cell(row=row, column=4, value=proyecto.get_estado_display())
        ws.cell(row=row, column=5, value=proyecto.get_prioridad_display())
        ws.cell(row=row, column=6, value=proyecto.fecha_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7, value=proyecto.fecha_fin_estimada.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=8, value=f"${proyecto.presupuesto_total:,.2f}")
        ws.cell(row=row, column=9, value=str(proyecto.responsable) if proyecto.responsable else "Sin asignar")
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'proyectos_sirius_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en response
    wb.save(response)
    return response

# ============ VISTA DE EXPORTACIÓN A PDF ============

@login_required
def exportar_proyectos_pdf(request):
    """Exportar proyectos a PDF"""
    proyectos = Proyecto.objects.all()[:20]
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="proyectos_sirius.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    data = [['Proyecto', 'Cliente', 'Estado', 'Presupuesto']]
    for p in proyectos:
        data.append([
            p.nombre[:30],
            str(p.cliente)[:25],
            p.get_estado_display(),
            f"${p.presupuesto_total:,.0f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

# ============ AJAX PARA CÁLCULOS AUTOMÁTICOS ============

@login_required
def calcular_total_presupuesto(request):
    """Calcular total del presupuesto via AJAX"""
    if request.method == 'POST':
        subtotal = float(request.POST.get('subtotal', 0))
        iva = float(request.POST.get('iva', 0))
        total = subtotal + iva
        
        return JsonResponse({'total': f'{total:.2f}'})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)